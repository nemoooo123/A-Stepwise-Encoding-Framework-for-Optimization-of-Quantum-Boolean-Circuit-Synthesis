"""
把 exp_5000_merged（AE-QTS 100 次）與 results/ 底下對應位元的
clamp / repair 100 次數據結合起來。

輸出：Combined_5000/
  {bit}bit_combined.xlsx   各方法的平均/標準差收斂曲線、100 次最終解、統計摘要
  {bit}bit_comparison.png  三種方法的收斂比較圖（含平均最佳解與 100 次最好解標註）
  overall_summary.xlsx     所有位元的總表
"""

import glob
import os

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

# --- Configuration ---
MERGED_ROOT = "exp_5000_merged"      # merge_5000.py 的輸出
RESULTS_ROOT = "results"             # 另外提供的 clamp / repair 數據
OUTPUT_DIR = "Combined_5000"
BITS = [6, 7, 8, 9, 10]
VARIANTS = ["clamp", "repair"]

# results 的 Run_xxx 工作表欄位（第 5 欄是 best-so-far，與 AE-QTS 收斂曲線同義）
RUN_COLS = ["iteration", "best_gc", "avg_gc", "worst_gc", "global_best_gc", "entropy"]
CURVE_COL = "global_best_gc"

METHOD_STYLE = {
    "AE-QTS":         {"color": "#F40E0E", "lw": 2.2, "ls": "-"},
    "AE-QTS-clamp":   {"color": "#1F77B4", "lw": 1.8, "ls": "--"},
    "AE-QTS-repair":  {"color": "#2CA02C", "lw": 1.8, "ls": "-."},
}


# ---------- 讀取 ----------

def load_merged(bit):
    """讀 exp_5000_merged 的 AE-QTS 100 次收斂曲線，回傳 (n_runs, n_gen) 陣列。"""
    pattern = os.path.join(MERGED_ROOT, f"{bit}_bit", "AE-QTS_Results", "*.xlsx")
    files = [f for f in glob.glob(pattern) if not os.path.basename(f).startswith("~$")]
    if not files:
        return None
    df = pd.read_excel(files[0], sheet_name="Total_Gate_Count", index_col=0)
    df.index = df.index.astype(str).str.strip()
    df = df[df.index.str.startswith("Trial_")]
    cols = [c for c in df.columns if str(c).startswith("Gen_")]
    return df[cols].apply(pd.to_numeric, errors="coerce").to_numpy(dtype=float)


def load_results_variant(bit, variant):
    """讀 results/{bit}bit 的某個 variant，回傳 (curves, summary_dict)。"""
    import openpyxl

    files = glob.glob(os.path.join(RESULTS_ROOT, f"{bit}bit", f"{bit}bit-{variant}-*.xlsx"))
    files = [f for f in files if not os.path.basename(f).startswith("~$")]
    if not files:
        return None, {}
    path = sorted(files)[-1]  # 有多個時取時間戳最新的

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    summary = {}
    if "Summary" in wb.sheetnames:
        for row in wb["Summary"].iter_rows(values_only=True):
            if row and len(row) > 1 and row[0] is not None:
                summary[str(row[0]).strip()] = row[1]

    col_idx = RUN_COLS.index(CURVE_COL)
    curves = []
    for name in wb.sheetnames:
        if not name.startswith("Run_"):
            continue
        ws = wb[name]
        vals = [r[col_idx] for r in ws.iter_rows(min_row=2, values_only=True) if r[0] is not None]
        curves.append([float(v) for v in vals])
    wb.close()

    if not curves:
        return None, summary
    width = min(len(c) for c in curves)
    return np.array([c[:width] for c in curves], dtype=float), summary


# ---------- 統計與繪圖 ----------

def stats_of(curves, time_sec=None):
    finals = curves[:, -1]
    return {
        "runs": len(curves),
        "generations": curves.shape[1],
        "mean_best": float(finals.mean()),
        "std_best": float(finals.std(ddof=0)),
        "best": float(finals.min()),
        "worst": float(finals.max()),
        "best_hit": int((finals == finals.min()).sum()),
        "median": float(np.median(finals)),
        "total_time_s": time_sec,
    }


def plot_comparison(bit, data, save_path):
    plt.figure(figsize=(12, 7))
    ax = plt.gca()

    for method, info in data.items():
        curves = info["curves"]
        style = METHOD_STYLE[method]
        mean = curves.mean(axis=0)
        std = curves.std(axis=0, ddof=0)
        gens = np.arange(1, len(mean) + 1)
        s = info["stats"]

        label = (f"{method}  (mean {s['mean_best']:.2f}±{s['std_best']:.2f}, "
                 f"best {s['best']:.0f}×{s['best_hit']})")
        ax.plot(gens, mean, color=style["color"], lw=style["lw"], ls=style["ls"], label=label)
        ax.fill_between(gens, mean - std, mean + std, color=style["color"], alpha=0.12)

        # 平均最佳解（實線水平參考線）與 100 次中最好的解（點線）
        ax.axhline(s["mean_best"], color=style["color"], lw=0.9, ls=":", alpha=0.75)
        ax.axhline(s["best"], color=style["color"], lw=0.9, ls=(0, (1, 3)), alpha=0.55)

    # 把標註集中在右側，依數值排序後錯開避免重疊
    last_gen = max(info["curves"].shape[1] for info in data.values())
    entries = sorted(data.items(), key=lambda kv: kv[1]["stats"]["mean_best"])
    for i, (method, info) in enumerate(entries):
        s = info["stats"]
        color = METHOD_STYLE[method]["color"]
        ax.annotate(f"{method}: mean {s['mean_best']:.2f} / best {s['best']:.0f}",
                    xy=(last_gen, s["mean_best"]),
                    xytext=(-10, 12 + 16 * i), textcoords="offset points",
                    ha="right", va="bottom", fontsize=9, fontweight="bold", color=color,
                    bbox=dict(boxstyle="round,pad=0.28", fc="white", ec=color, alpha=0.85))

    ax.set_title(f"{bit}-bit Convergence Comparison (100 runs each, 5000 generations)",
                 fontsize=14, fontweight="bold")
    ax.set_xlabel("Generation", fontsize=12)
    ax.set_ylabel("Gate Count (best-so-far)", fontsize=12)
    ax.grid(True, linestyle="--", alpha=0.6)
    ax.legend(loc="upper right", frameon=True, shadow=True, prop={"size": 9})
    plt.tight_layout()
    plt.savefig(save_path, dpi=150)
    plt.close()


# ---------- 主流程 ----------

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    overall = []

    for bit in BITS:
        print(f"\n=== {bit}-bit ===")
        data = {}

        merged = load_merged(bit)
        if merged is not None:
            data["AE-QTS"] = {"curves": merged, "stats": stats_of(merged), "summary": {}}
            print(f"  AE-QTS         : {merged.shape[0]} runs x {merged.shape[1]} gens")
        else:
            print(f"  [warn] 找不到 {MERGED_ROOT}/{bit}_bit 的合併資料")

        for variant in VARIANTS:
            curves, summary = load_results_variant(bit, variant)
            if curves is None:
                print(f"  [warn] results/{bit}bit 找不到 {variant}")
                continue
            method = f"AE-QTS-{variant}"
            data[method] = {
                "curves": curves,
                "stats": stats_of(curves, summary.get("Total seconds")),
                "summary": summary,
            }
            print(f"  {method:<15}: {curves.shape[0]} runs x {curves.shape[1]} gens")

        if not data:
            continue

        # --- 合併後的 xlsx ---
        xlsx_path = os.path.join(OUTPUT_DIR, f"{bit}bit_combined.xlsx")
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            # 1) 統計摘要
            summary_df = pd.DataFrame({m: info["stats"] for m, info in data.items()}).T
            summary_df.index.name = "Method"
            summary_df.to_excel(writer, sheet_name="Summary")

            # 2) 平均 / 標準差收斂曲線（每列一個方法）
            rows, index = [], []
            for method, info in data.items():
                c = info["curves"]
                rows.append(c.mean(axis=0)); index.append(f"{method}_Mean")
                rows.append(c.std(axis=0, ddof=0)); index.append(f"{method}_Std")
                rows.append(c.min(axis=0)); index.append(f"{method}_Best")
            width = max(len(r) for r in rows)
            padded = [np.pad(r, (0, width - len(r)), constant_values=np.nan) for r in rows]
            pd.DataFrame(padded, index=index,
                         columns=[f"Gen_{i + 1}" for i in range(width)]
                         ).to_excel(writer, sheet_name="Mean_Curves")

            # 3) 每個方法 100 次的最終解
            finals = pd.DataFrame(
                {m: pd.Series(info["curves"][:, -1]) for m, info in data.items()}
            )
            finals.index = [f"Run_{i + 1}" for i in range(len(finals))]
            finals.to_excel(writer, sheet_name="Final_Solutions")

            # 4) 原始 results Summary 欄位（問題設定、permutation 等）
            meta = {m: info["summary"] for m, info in data.items() if info["summary"]}
            if meta:
                pd.DataFrame(meta).to_excel(writer, sheet_name="Problem_Setup")
        print(f"  -> {xlsx_path}")

        # --- 比較圖 ---
        png_path = os.path.join(OUTPUT_DIR, f"{bit}bit_comparison.png")
        plot_comparison(bit, data, png_path)
        print(f"  -> {png_path}")

        for method, info in data.items():
            overall.append({"bit": bit, "method": method, **info["stats"]})

    if overall:
        df = pd.DataFrame(overall).set_index(["bit", "method"])
        out = os.path.join(OUTPUT_DIR, "overall_summary.xlsx")
        df.to_excel(out)
        print(f"\n[System] 總表：{out}")
        print(df.to_string())


if __name__ == "__main__":
    main()
